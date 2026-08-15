"""Create a formatted all-in-one Excel workbook for manual labeling.

This workbook includes:
- dataset index
- compact manual labeling worksheet
- blank labeling template
- taxi-only review sheet
- the five full raw datasets as formatted tabs
- a compact codebook
"""

from __future__ import annotations

import csv
from datetime import datetime
import os
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
BASE = Path(os.environ.get("BUCKAROO_MANUAL_LABEL_DIR", ROOT / "outputs" / "manual_labeling_5_datasets"))
OUTPUT = BASE / "manual_labeling_with_all_datasets_formatted.xlsx"

INDEX_CSV = BASE / "manual_labeling_5_datasets_index.csv"
FILLED_CSV = BASE / "manual_column_labeling_professional_with_taxi_filled.csv"
BLANK_CSV = BASE / "manual_column_labeling_professional_blank.csv"

DATASETS = [
    ("taxi_trips", BASE / "taxi_trips.csv"),
    ("us_airports", BASE / "us_airports.csv"),
    ("stock_prices", BASE / "stock_prices.csv"),
    ("adult_census_income", BASE / "adult_census_income.csv"),
    ("diamonds_pricing", BASE / "diamonds_pricing.csv"),
]

INK = "12343B"
TEAL = "0F8B8D"
PALE = "F6FAFC"
CREAM = "FFFDF6"
LINE = "DDEAF0"
WHITE = "FFFFFF"
AMBER = "F6E7BE"

header_fill = PatternFill("solid", fgColor=INK)
evidence_fill = PatternFill("solid", fgColor=PALE)
editable_fill = PatternFill("solid", fgColor=CREAM)
warning_fill = PatternFill("solid", fgColor=AMBER)
white_font = Font(name="Aptos", size=10, bold=True, color=WHITE)
body_font = Font(name="Aptos", size=10, color="1F2933")
small_font = Font(name="Aptos", size=9, color="1F2933")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_alignment = Alignment(vertical="top", wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin_bottom = Border(bottom=Side(style="thin", color=LINE))


def read_csv(path: Path) -> List[List[str]]:
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        return list(csv.reader(handle))


def parse_value(dataset_id: str, header: str, value: str):
    if value == "":
        return None

    date_columns = {
        "taxi_trips": {"pickup", "dropoff"},
        "stock_prices": {"date"},
    }
    text_columns = {
        "taxi_trips": {"color", "payment", "pickup_zone", "dropoff_zone", "pickup_borough", "dropoff_borough"},
        "us_airports": {"iata", "name", "city", "state", "country"},
        "stock_prices": {"symbol"},
        "adult_census_income": {
            "workclass",
            "education",
            "marital.status",
            "occupation",
            "relationship",
            "race",
            "sex",
            "native.country",
            "income",
        },
        "diamonds_pricing": {"cut", "color", "clarity"},
    }

    if header in date_columns.get(dataset_id, set()):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass
        return value

    if header not in text_columns.get(dataset_id, set()):
        try:
            if value.replace(".", "", 1).replace("-", "", 1).isdigit():
                return float(value) if "." in value else int(value)
        except ValueError:
            return value

    return value


def write_rows(ws, rows: Iterable[Iterable], *, dataset_id: str | None = None) -> None:
    headers = None
    for row_index, row in enumerate(rows, start=1):
        if row_index == 1 or dataset_id is None:
            headers = list(row)
            ws.append(headers)
            continue
        ws.append([parse_value(dataset_id, headers[idx], value) for idx, value in enumerate(row)])


def professional_width(header: str) -> int:
    small = {
        "row_count",
        "null_ratio",
        "unique_ratio",
        "is_primary_key",
        "is_foreign_key",
        "has_missing_values",
        "sbert_use_recommended",
    }
    medium = {
        "dataset_id",
        "column_name",
        "manual_true_role",
        "manual_secondary_role",
        "manual_physical_type",
        "semantic_group",
        "manual_label_confidence",
        "requires_semantic_ml",
        "simple_rules_enough",
        "adaptive_sampling_priority",
        "expected_buckaroo_role",
        "expected_warning_type",
    }
    large = {
        "sample_values",
        "common_values",
        "expected_candidate_roles",
        "expected_confidence_behavior",
        "advanced_ml_analysis_reason",
        "ui_user_facing_explanation",
        "profiler_failure_mode_to_test",
        "professor_question_to_answer",
        "paper_claim_supported",
        "why_this_label",
        "edge_case_or_risk",
        "reviewer_notes",
    }
    if header in small:
        return 14
    if header in medium:
        return 22
    if header in large:
        return 48
    return 24


def raw_width(dataset_id: str, header: str) -> int:
    if header in {"pickup", "dropoff"}:
        return 23
    if header == "date":
        return 14
    if header in {"pickup_zone", "dropoff_zone", "name", "city", "occupation", "native.country", "education"}:
        return 30
    if dataset_id == "diamonds_pricing":
        return 13
    if dataset_id == "adult_census_income":
        return max(14, min(24, len(header) + 4))
    return max(12, min(24, len(header) + 4))


def style_common(ws, *, freeze: str, row_height: int = 42) -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = header_alignment
        cell.border = thin_bottom
    ws.row_dimensions[1].height = 44

    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = row_height
        for cell in row:
            cell.font = body_font
            cell.alignment = body_alignment
            cell.border = thin_bottom


def style_professional(ws) -> None:
    # Freeze only the header row. Freezing columns A:B looked like duplicate
    # columns in Excel's split-pane view and confused manual labeling.
    style_common(ws, freeze="A2", row_height=86)
    headers = [cell.value for cell in ws[1]]
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = professional_width(str(header))
        fill = evidence_fill if col_idx <= 7 else editable_fill
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
            for c in cell:
                c.fill = fill
                if str(header).endswith("_ratio"):
                    c.number_format = "0.000"
                elif header == "row_count":
                    c.number_format = "#,##0"

    validation_values = {
        "manual_true_role": "datetime,numeric_measure,location_name,categorical,geographic_coordinate,entity_name,identifier_code,boolean,text,unknown_or_mixed",
        "is_primary_key": "yes,no,maybe",
        "is_foreign_key": "yes,no,maybe",
        "could_be_key_by_uniqueness": "yes,no,maybe",
        "should_be_key_candidate_for_buckaroo": "yes,no,maybe",
        "is_high_uniqueness_but_not_key": "yes,no,maybe",
        "requires_semantic_ml": "yes,no,maybe",
        "sbert_use_recommended": "yes,no,maybe",
        "simple_rules_enough": "yes,no,maybe",
        "adaptive_sampling_priority": "low,medium,high",
        "manual_label_confidence": "low,medium,high",
        "should_buckaroo_warn": "yes,no,optional",
    }
    for col_idx, header in enumerate(headers, start=1):
        if header not in validation_values:
            continue
        dv = DataValidation(type="list", formula1=f'"{validation_values[header]}"', allow_blank=True)
        ws.add_data_validation(dv)
        col = get_column_letter(col_idx)
        dv.add(f"{col}2:{col}{ws.max_row}")


def style_raw(ws, dataset_id: str) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.sheet_format.defaultRowHeight = 24

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = header_alignment
        cell.border = thin_bottom
    ws.row_dimensions[1].height = 44

    headers = [cell.value for cell in ws[1]]
    for col_idx, header in enumerate(headers, start=1):
        col = get_column_letter(col_idx)
        ws.column_dimensions[col].width = raw_width(dataset_id, str(header))

        # Keep the expensive full-sheet style pass out of large raw tabs.
        # Date/time formats matter most for readability and prevent Excel #######.
        if header in {"pickup", "dropoff", "date"}:
            fmt = "yyyy-mm-dd hh:mm:ss" if header in {"pickup", "dropoff"} else "yyyy-mm-dd"
            for row_idx in range(2, ws.max_row + 1):
                ws[f"{col}{row_idx}"].number_format = fmt


def style_index(ws) -> None:
    style_common(ws, freeze="A2", row_height=64)
    for idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value)
        width = 52 if header in {"column_names", "source_url", "local_path", "copied_csv"} else 24
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = editable_fill


def add_codebook(wb: Workbook) -> None:
    ws = wb.create_sheet("Codebook")
    rows = [
        ("Field", "How to use it"),
        ("manual_true_role", "Human semantic label: datetime, numeric_measure, location_name, categorical, etc."),
        ("is_primary_key", "Only yes if this column truly identifies each row."),
        ("could_be_key_by_uniqueness", "Yes if statistics make it look key-like. This is not the same as being a real key."),
        ("should_be_key_candidate_for_buckaroo", "Yes only if Buckaroo should actually consider it as a key."),
        ("is_high_uniqueness_but_not_key", "Main false-key flag for timestamps, locations, prices, names, and codes that look unique but are not row identity."),
        ("requires_semantic_ml", "Use yes/maybe when rules and statistics are not enough."),
        ("sbert_use_recommended", "Use yes/maybe for place names, occupations, descriptions, product names, organizations, and ambiguous text."),
        ("adaptive_sampling_priority", "High means Buckaroo should sample more rows before deciding."),
        ("expected_buckaroo_role", "What the improved profiler should output."),
        ("expected_warning_type", "Warning Buckaroo should show to the user, if any."),
        ("profiler_failure_mode_to_test", "The exact kind of profiler mistake this column helps test."),
    ]
    for row in rows:
        ws.append(row)
    style_common(ws, freeze="A2", row_height=56)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = editable_fill


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Dataset Index")
    write_rows(ws, read_csv(INDEX_CSV))
    style_index(ws)

    ws = wb.create_sheet("Labeling Worksheet")
    write_rows(ws, read_csv(FILLED_CSV))
    style_professional(ws)

    ws = wb.create_sheet("Blank Template")
    write_rows(ws, read_csv(BLANK_CSV))
    style_professional(ws)

    filled_rows = read_csv(FILLED_CSV)
    taxi_rows = [filled_rows[0]] + [row for row in filled_rows[1:] if row and row[0] == "taxi_trips"]
    ws = wb.create_sheet("Taxi Review")
    write_rows(ws, taxi_rows)
    style_professional(ws)

    for dataset_id, csv_path in DATASETS:
        ws = wb.create_sheet(dataset_id)
        write_rows(ws, read_csv(csv_path), dataset_id=dataset_id)
        style_raw(ws, dataset_id)

    add_codebook(wb)

    wb.save(OUTPUT)

    # Lightweight verification pass.
    check = load_workbook(OUTPUT, read_only=True, data_only=True)
    print(OUTPUT)
    print(f"sheets={len(check.sheetnames)}")
    for name in check.sheetnames:
        ws = check[name]
        print(f"{name}: rows={ws.max_row} cols={ws.max_column}")
    check.close()


if __name__ == "__main__":
    main()
