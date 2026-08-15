"""Create a formatted workbook for reviewing generated human labels.

This is the practical review file: every selected dataset column receives a
generated human semantic label, and the reviewer only needs to accept, correct,
or flag uncertainty.
"""

from __future__ import annotations

import csv
from collections import Counter
from datetime import datetime
import os
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
BASE = Path(os.environ.get("BUCKAROO_MANUAL_LABEL_DIR", ROOT / "outputs" / "manual_labeling_5_datasets"))
OUTPUT = BASE / "manual_labeling_peer_review_final.xlsx"
FILLED_OUTPUT = BASE / "manual_labeling_peer_review_final_filled_no_blanks.xlsx"
GENERATED_CSV = BASE / "manual_labeling_peer_review_final.csv"
REVIEW_GUIDE = BASE / "manual_labeling_review_guide_easy.md"

DATASETS = [
    ("taxi_trips", BASE / "taxi_trips.csv"),
    ("us_airports", BASE / "us_airports.csv"),
    ("stock_prices", BASE / "stock_prices.csv"),
    ("adult_census_income", BASE / "adult_census_income.csv"),
    ("diamonds_pricing", BASE / "diamonds_pricing.csv"),
]

DATASET_CONTEXT = {
    "taxi_trips": (
        "Urban mobility / transactions",
        "Has timestamps, money, counts, and locations. Good for testing false keys from unique-looking datetime columns.",
    ),
    "us_airports": (
        "Geography / reference data",
        "Has natural codes, names, city/state/country fields, and latitude/longitude. Good for testing geography safeguards.",
    ),
    "stock_prices": (
        "Finance / time series",
        "Has an entity code plus date pattern. Good for testing composite-key thinking instead of single-column key guessing.",
    ),
    "adult_census_income": (
        "Demographics / social data",
        "Has sensitive attributes, numeric-coded categories, missingness, and mixed categorical semantics.",
    ),
    "diamonds_pricing": (
        "Retail / product measurements",
        "Has product quality categories, physical measurements, and price. Good for separating measures from categories.",
    ),
}

REVIEW_GUIDE_SECTIONS = [
    (
        "What You Are Doing",
        "You are not inventing labels from scratch. You are reviewing generated human labels. For each row, ask: does this label match what the column means in real life? If yes, mark review_status as accepted. If no, change the corrected fields and explain why.",
    ),
    (
        "Fast Review Steps",
        "1. Look at dataset_id and column_name. 2. Read sample_values and common_values. 3. Check manual_true_role. 4. Check is_primary_key and is_high_uniqueness_but_not_key. 5. Check expected_buckaroo_role. 6. Mark review_status as accepted, change_needed, or unsure.",
    ),
    (
        "Accepted",
        "Use accepted when the generated label matches the real-world meaning of the column. Example: taxi fare is numeric_measure and money; taxi pickup is datetime; diamond cut is ordinal_category.",
    ),
    (
        "Change Needed",
        "Use change_needed when the generated label is clearly wrong. Example: if a column labeled numeric_measure is actually an ID code, or a column labeled primary key is just a timestamp, coordinate, price, or name.",
    ),
    (
        "Unsure",
        "Use unsure when you cannot confidently decide from the column name and values. This is not a failure. It means the benchmark needs human/domain review. Add a short note explaining what confused you.",
    ),
    (
        "manual_true_role",
        "This is the main human meaning label. datetime means date/time. numeric_measure means a number that measures something. categorical means repeated category. ordinal_category means category with order. location_name means place name. geographic_coordinate means latitude/longitude. identifier_code means code like stock ticker or airport code. entity_name means a real-world name.",
    ),
    (
        "Primary Key Rule",
        "A primary key must identify one row because that is its purpose, not just because it happens to be unique. Unique timestamps, prices, coordinates, names, and measurements should usually NOT be primary keys.",
    ),
    (
        "High Uniqueness But Not Key",
        "Mark this idea as correct when a column has many unique values but does not represent row identity. Good examples: pickup timestamp, dropoff timestamp, latitude, longitude. These are the exact false-key cases Buckaroo must avoid.",
    ),
    (
        "Identifier Code",
        "Use identifier_code for compact codes that identify an entity, like airport IATA code or stock ticker. But if the code repeats across time, like stock symbol, it is not a single-column primary key; it may be part of a composite key with date.",
    ),
    (
        "Composite Key",
        "A composite key means two or more columns together identify a row. For stock prices, symbol alone repeats and date alone repeats, but symbol + date can identify one stock-price observation.",
    ),
    (
        "Numeric Measure",
        "Use numeric_measure for numbers that measure amount, quantity, price, distance, age, hours, dimensions, etc. These are usually not keys. They may need range checks, unit checks, and outlier warnings.",
    ),
    (
        "Categorical vs Ordinal",
        "categorical means categories with no order, like payment method. ordinal_category means categories with meaningful order, like education level, diamond cut, diamond color grade, or diamond clarity.",
    ),
    (
        "Location Columns",
        "Location columns include city, state, country, borough, zone, latitude, and longitude. These should usually not become primary keys. Latitude/longitude can look highly unique, but they describe position, not row identity.",
    ),
    (
        "Sensitive Columns",
        "Sensitive columns include race, sex, income, relationship, marital status, and similar demographic fields. The label can be correct, but Buckaroo should warn or treat them carefully in UI/research claims.",
    ),
    (
        "SBERT / ML Decision",
        "Use SBERT or advanced semantic ML only when simple rules are not enough. For clear columns like price, latitude, pickup, fare, rules are enough. For vague names, occupation labels, entity names, or location names, SBERT/LLM may help explain meaning.",
    ),
    (
        "Metanome Comparison",
        "Metanome finds mathematical uniqueness and dependencies. It can say a timestamp or coordinate is unique, but that does not mean it is a real primary key. Your human label tells Buckaroo when to reject mathematically tempting false keys.",
    ),
    (
        "Deequ Comparison",
        "Deequ is useful for quality checks: completeness, distinctness, allowed values, min/max, and distributions. It does not fully understand semantics by itself, so your labels explain what checks matter.",
    ),
    (
        "DataProfiler Comparison",
        "DataProfiler helps with physical type, nulls, uniqueness, and sometimes semantic labels. Your human labels are the benchmark for deciding whether those inferred labels actually make sense.",
    ),
    (
        "What To Edit",
        "Most of the time, only edit review_status. If wrong, edit corrected_manual_true_role, corrected_is_primary_key, and corrected_notes. Do not waste time rewriting every explanation field unless something is clearly wrong.",
    ),
    (
        "What Counts As A Good Review",
        "A good review says accepted for rows that match the real-world meaning, change_needed for rows where the semantic label or key label is wrong, and unsure for ambiguous rows. The goal is correctness, not perfection theater.",
    ),
]

FIELDS = [
    "review_status",
    "corrected_manual_true_role",
    "corrected_is_primary_key",
    "corrected_notes",
    "dataset_id",
    "column_name",
    "row_count",
    "null_ratio",
    "unique_ratio",
    "sample_values",
    "common_values",
    "non_null_count",
    "null_count",
    "unique_count",
    "duplicate_value_count",
    "top_value",
    "top_value_count",
    "top_value_ratio",
    "manual_true_role",
    "manual_secondary_role",
    "manual_physical_type",
    "semantic_group",
    "semantic_subtype",
    "unit_or_currency",
    "temporal_granularity",
    "geographic_level",
    "entity_type",
    "domain_vocabulary",
    "generated_label_confidence",
    "is_primary_key",
    "is_foreign_key",
    "could_be_key_by_uniqueness",
    "should_be_key_candidate_for_buckaroo",
    "is_high_uniqueness_but_not_key",
    "key_rejection_reason",
    "is_identifier_like_code",
    "identifier_code_type",
    "entity_name_vs_identifier",
    "is_surrogate_key",
    "is_natural_key",
    "is_composite_key_part",
    "possible_composite_key_with",
    "foreign_key_target_if_known",
    "is_datetime_or_lifecycle_event",
    "lifecycle_event_type",
    "is_geographic_or_location",
    "location_semantic_type",
    "coordinate_pair_partner",
    "is_measure_or_metric",
    "is_money_amount",
    "is_count_or_quantity",
    "is_rate_ratio_or_percentage",
    "is_bounded_numeric",
    "expected_min",
    "expected_max",
    "zero_heavy",
    "negative_allowed",
    "outlier_sensitive",
    "nominal_category",
    "ordinal_category",
    "boolean_like",
    "free_text_or_description",
    "has_missing_values",
    "missingness_severity",
    "missingness_semantics",
    "invalid_value_risk",
    "standardization_risk",
    "mixed_type_risk",
    "format_consistency_risk",
    "unit_consistency_risk",
    "sensitive_or_pii_risk",
    "pii_type",
    "ucc_candidate_status",
    "fd_determinant_candidate",
    "fd_dependent_candidate",
    "ind_candidate_status",
    "order_dependency_candidate",
    "derived_or_calculated_field",
    "derived_from_columns",
    "sample_size_sensitivity",
    "small_sample_false_key_risk",
    "noise_sensitivity",
    "requires_semantic_ml",
    "recommended_semantic_model",
    "sbert_use_recommended",
    "simple_rules_enough",
    "advanced_ml_analysis_reason",
    "adaptive_sampling_priority",
    "min_recommended_sample_size",
    "confidence_interval_priority",
    "expected_candidate_roles",
    "expected_candidate_confidence_pattern",
    "low_confidence_adaptive_sampling_trigger",
    "adaptive_sampling_stop_condition",
    "expected_buckaroo_role",
    "expected_warning_type",
    "should_buckaroo_warn",
    "ui_should_show_confidence_interval",
    "ui_should_show_warning_badge",
    "ui_user_facing_explanation",
    "profiler_failure_mode_to_test",
    "metanome_expected_behavior",
    "deequ_expected_constraint",
    "dataprofiler_expected_behavior",
    "llm_semantic_label_expected",
    "professor_question_to_answer",
    "paper_claim_supported",
    "benchmark_importance",
    "manual_label_priority",
    "poster_worthy_example",
    "why_this_label",
    "edge_case_or_risk",
    "reviewer_notes",
]

INK = "12343B"
PALE = "F6FAFC"
CREAM = "FFFDF6"
GREEN = "E8F6EF"
AMBER = "FFF3CD"
LINE = "DDEAF0"
WHITE = "FFFFFF"

header_fill = PatternFill("solid", fgColor=INK)
review_fill = PatternFill("solid", fgColor=GREEN)
evidence_fill = PatternFill("solid", fgColor=PALE)
label_fill = PatternFill("solid", fgColor=CREAM)
warning_fill = PatternFill("solid", fgColor=AMBER)
header_font = Font(name="Aptos", size=10, bold=True, color=WHITE)
body_font = Font(name="Aptos", size=10, color="1F2933")
small_font = Font(name="Aptos", size=9, color="1F2933")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_alignment = Alignment(vertical="top", wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin_bottom = Border(bottom=Side(style="thin", color=LINE))


def read_csv(path: Path) -> List[List[str]]:
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        return list(csv.reader(handle))


def ratio(num: int, den: int) -> str:
    return f"{num / den:.6f}" if den else ""


def dataset_stats(dataset_id: str, path: Path) -> List[Dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        reader = csv.DictReader(handle)
        columns = reader.fieldnames or []
        stats = {col: {"nulls": 0, "unique": set(), "examples": [], "counter": Counter()} for col in columns}
        total = 0
        for record in reader:
            total += 1
            for col in columns:
                value = (record.get(col) or "").strip()
                if value == "":
                    stats[col]["nulls"] += 1
                    continue
                stats[col]["unique"].add(value)
                if len(stats[col]["examples"]) < 5 and value not in stats[col]["examples"]:
                    stats[col]["examples"].append(value)
                stats[col]["counter"][value] += 1

    rows: List[Dict[str, str]] = []
    for col in columns:
        non_null = total - int(stats[col]["nulls"])
        unique_count = len(stats[col]["unique"])
        duplicate_value_count = max(non_null - unique_count, 0)
        top_value = ""
        top_value_count = 0
        if stats[col]["counter"]:
            top_value, top_value_count = stats[col]["counter"].most_common(1)[0]
        rows.append(
            {
                "dataset_id": dataset_id,
                "column_name": col,
                "row_count": str(total),
                "null_ratio": ratio(int(stats[col]["nulls"]), total),
                "unique_ratio": ratio(unique_count, non_null),
                "sample_values": " | ".join(stats[col]["examples"]),
                "common_values": " | ".join(f"{v} ({n})" for v, n in stats[col]["counter"].most_common(5)),
                "non_null_count": str(non_null),
                "null_count": str(int(stats[col]["nulls"])),
                "unique_count": str(unique_count),
                "duplicate_value_count": str(duplicate_value_count),
                "top_value": str(top_value),
                "top_value_count": str(top_value_count),
                "top_value_ratio": ratio(top_value_count, non_null),
            }
        )
    return rows


def base_label(**overrides: str) -> Dict[str, str]:
    row = {field: "" for field in FIELDS}
    row.update(
        {
            "review_status": "needs_review",
            "generated_label_confidence": "medium",
            "is_primary_key": "no",
            "is_foreign_key": "no",
            "could_be_key_by_uniqueness": "no",
            "should_be_key_candidate_for_buckaroo": "no",
            "is_high_uniqueness_but_not_key": "no",
            "is_datetime_or_lifecycle_event": "no",
            "is_geographic_or_location": "no",
            "is_measure_or_metric": "no",
            "is_money_amount": "no",
            "is_count_or_quantity": "no",
            "nominal_category": "no",
            "ordinal_category": "no",
            "has_missing_values": "no",
            "missingness_severity": "none",
            "sensitive_or_pii_risk": "no",
            "requires_semantic_ml": "no",
            "recommended_semantic_model": "rules_and_statistics",
            "sbert_use_recommended": "no",
            "simple_rules_enough": "yes",
            "adaptive_sampling_priority": "low",
            "should_buckaroo_warn": "no",
        }
    )
    row.update(overrides)
    return row


def money_measure(name: str, subtype: str, failure: str, reason: str, *, zero: bool = False) -> Dict[str, str]:
    return base_label(
        manual_true_role="numeric_measure",
        manual_secondary_role=subtype,
        manual_physical_type="float",
        semantic_group="measure",
        generated_label_confidence="high",
        is_measure_or_metric="yes",
        is_money_amount="yes",
        expected_candidate_roles="numeric_measure:high; categorical:low_or_medium; primary_key:low",
        expected_buckaroo_role="numeric_measure_money",
        expected_warning_type="zero_heavy_distribution_optional" if zero else "range_outlier_optional",
        should_buckaroo_warn="optional",
        profiler_failure_mode_to_test=failure,
        professor_question_to_answer="Why is a numeric amount not an identifier even if it has many distinct values?",
        paper_claim_supported="Buckaroo should separate numeric measures from identifiers and expose confidence/warnings.",
        why_this_label=reason,
        edge_case_or_risk="Money columns can repeat, be zero-heavy, or have many distinct decimals, but they are measures, not row identity.",
    )


def numeric_measure(subtype: str, physical: str, failure: str, reason: str, *, count: bool = False) -> Dict[str, str]:
    return base_label(
        manual_true_role="numeric_measure",
        manual_secondary_role=subtype,
        manual_physical_type=physical,
        semantic_group="measure",
        generated_label_confidence="high",
        is_measure_or_metric="yes",
        is_count_or_quantity="yes" if count else "no",
        expected_candidate_roles="numeric_measure:high; categorical:medium_if_low_cardinality; primary_key:low",
        expected_buckaroo_role="numeric_measure_count" if count else "numeric_measure",
        expected_warning_type="low_cardinality_numeric_optional" if count else "range_outlier_optional",
        should_buckaroo_warn="optional" if count else "no",
        profiler_failure_mode_to_test=failure,
        professor_question_to_answer="Can Buckaroo avoid confusing low-cardinality numbers with plain categories?",
        paper_claim_supported="Buckaroo should report candidate roles when statistical and semantic evidence differ.",
        why_this_label=reason,
        edge_case_or_risk="A numeric column may look categorical when it has few distinct values, but its meaning can still be a measurement.",
    )


def categorical(subtype: str, reason: str, *, ordinal: bool = False, sensitive: bool = False, ml: str = "no") -> Dict[str, str]:
    return base_label(
        manual_true_role="categorical" if not ordinal else "ordinal_category",
        manual_secondary_role=subtype,
        manual_physical_type="string",
        semantic_group="category",
        generated_label_confidence="high",
        nominal_category="no" if ordinal else "yes",
        ordinal_category="yes" if ordinal else "no",
        sensitive_or_pii_risk="yes" if sensitive else "no",
        requires_semantic_ml=ml,
        recommended_semantic_model="SBERT_or_LLM_for_text_semantics" if ml in {"yes", "maybe"} else "dictionary_lookup_or_value_set_rules",
        sbert_use_recommended=ml,
        simple_rules_enough="maybe" if ml in {"yes", "maybe"} else "yes",
        expected_candidate_roles=("ordinal_category:high; categorical:medium; primary_key:low" if ordinal else "categorical:high; primary_key:low"),
        expected_buckaroo_role="ordinal_category" if ordinal else "categorical",
        expected_warning_type="sensitive_attribute_optional" if sensitive else "none",
        should_buckaroo_warn="optional" if sensitive else "no",
        profiler_failure_mode_to_test="semantic_category_detection",
        professor_question_to_answer="Can Buckaroo distinguish semantic categories from IDs, text, and measures?",
        paper_claim_supported="Semantic profiling needs both statistics and column/value meaning.",
        why_this_label=reason,
        edge_case_or_risk="Categorical values can need semantic interpretation, especially when labels are domain-specific.",
    )


def location(subtype: str, reason: str, *, coordinate: bool = False, code: bool = False, key_like: bool = False) -> Dict[str, str]:
    role = "geographic_coordinate" if coordinate else ("identifier_code" if code else "location_name")
    return base_label(
        manual_true_role=role,
        manual_secondary_role=subtype,
        manual_physical_type="float" if coordinate else "string",
        semantic_group="location",
        generated_label_confidence="high",
        is_geographic_or_location="yes",
        is_measure_or_metric="yes" if coordinate else "no",
        is_identifier_like_code="yes" if code else "no",
        could_be_key_by_uniqueness="yes" if key_like else "maybe" if coordinate else "no",
        should_be_key_candidate_for_buckaroo="yes" if code and key_like else "no",
        is_high_uniqueness_but_not_key="yes" if coordinate else "no",
        key_rejection_reason="coordinate_or_location_not_row_identity" if coordinate else "location_field_not_row_identity",
        requires_semantic_ml="no" if coordinate or code else "maybe",
        recommended_semantic_model="coordinate_range_rules" if coordinate else "location_dictionary_or_SBERT",
        sbert_use_recommended="no" if coordinate or code else "maybe",
        simple_rules_enough="yes" if coordinate or code else "maybe",
        adaptive_sampling_priority="medium" if key_like or coordinate else "low",
        expected_candidate_roles=f"{role}:high; primary_key:{'medium' if code and key_like else 'low'}",
        expected_buckaroo_role=role if not coordinate else "geographic_coordinate_high_uniqueness",
        expected_warning_type="location_or_coordinate_not_primary_key" if coordinate else "location_field_not_key",
        should_buckaroo_warn="yes" if coordinate else "optional",
        profiler_failure_mode_to_test="geography_false_key_risk" if coordinate else "location_semantic_detection",
        professor_question_to_answer="Can Buckaroo prevent geography/location fields from being misclassified as primary keys?",
        paper_claim_supported="Geography-aware safeguards reduce false key behavior in semantic profiling.",
        why_this_label=reason,
        edge_case_or_risk="Location fields can look key-like or category-like, but they describe place semantics.",
    )


def datetime_label(subtype: str, reason: str, *, key_like: bool = False, composite: bool = False) -> Dict[str, str]:
    return base_label(
        manual_true_role="datetime",
        manual_secondary_role=subtype,
        manual_physical_type="datetime_string",
        semantic_group="temporal",
        generated_label_confidence="high",
        could_be_key_by_uniqueness="yes" if key_like else "no",
        should_be_key_candidate_for_buckaroo="no",
        is_high_uniqueness_but_not_key="yes" if key_like else "no",
        key_rejection_reason="timestamp_lifecycle_field_not_identity" if key_like else "date_field_not_single_row_identity",
        is_datetime_or_lifecycle_event="yes",
        requires_semantic_ml="no",
        recommended_semantic_model="datetime_parser_plus_column_name_rules",
        sbert_use_recommended="no",
        simple_rules_enough="yes",
        adaptive_sampling_priority="high" if key_like else "medium" if composite else "low",
        expected_candidate_roles="datetime:high; primary_key:low",
        expected_buckaroo_role="datetime_high_uniqueness" if key_like else "datetime",
        expected_warning_type="unique_timestamp_not_primary_key" if key_like else "none",
        should_buckaroo_warn="yes" if key_like else "no",
        profiler_failure_mode_to_test="false_key_from_high_uniqueness_datetime" if key_like else "datetime_role_detection",
        professor_question_to_answer="Why is a unique-looking date/time not automatically a primary key?",
        paper_claim_supported="Confidence-aware profiling separates timestamp uniqueness from identity.",
        why_this_label=reason,
        edge_case_or_risk="Datetime fields can look unique in samples but usually describe events or periods.",
    )


def labels_for(dataset_id: str, column: str) -> Dict[str, str]:
    taxi = {
        "pickup": datetime_label("event_start_time", "Taxi pickup timestamp; describes when the trip started.", key_like=True),
        "dropoff": datetime_label("event_end_time", "Taxi dropoff timestamp; describes when the trip ended.", key_like=True),
        "passengers": numeric_measure("discrete_count", "integer", "numeric_count_vs_category", "Passenger count is a whole-number quantity.", count=True),
        "distance": numeric_measure("continuous_measure", "float", "standard_numeric_measure_with_outliers", "Trip distance is a continuous numeric measurement."),
        "fare": money_measure("fare", "base_fare", "standard_money_measure", "Base fare is a money amount."),
        "tip": money_measure("tip", "tip_amount", "zero_heavy_money_measure", "Tip is a money amount with many zero values.", zero=True),
        "tolls": money_measure("tolls", "toll_amount", "low_cardinality_money_measure", "Tolls are money amounts with many zero values.", zero=True),
        "total": money_measure("total", "total_paid", "derived_total_measure_not_key", "Total payment is a derived-looking money amount."),
        "color": categorical("vehicle_type", "Taxi color/type is a repeated nominal category."),
        "payment": categorical("payment_method", "Payment method is a repeated nominal category."),
        "pickup_zone": location("pickup_zone_name", "Pickup zone is a named location; it may link to a zone lookup but is not row identity."),
        "dropoff_zone": location("dropoff_zone_name", "Dropoff zone is a named location; it may link to a zone lookup but is not row identity."),
        "pickup_borough": location("pickup_borough_name", "Pickup borough is a repeated geographic place label."),
        "dropoff_borough": location("dropoff_borough_name", "Dropoff borough is a repeated geographic place label."),
    }
    airports = {
        "iata": location("airport_code", "IATA/airport code identifies an airport in this reference table.", code=True, key_like=True),
        "name": base_label(
            manual_true_role="entity_name",
            manual_secondary_role="airport_name",
            manual_physical_type="string",
            semantic_group="entity",
            generated_label_confidence="medium",
            could_be_key_by_uniqueness="maybe",
            should_be_key_candidate_for_buckaroo="maybe",
            requires_semantic_ml="maybe",
            recommended_semantic_model="dictionary_lookup_plus_SBERT_or_LLM",
            sbert_use_recommended="maybe",
            simple_rules_enough="maybe",
            expected_candidate_roles="entity_name:high; primary_key:medium_if_unique; location_name:medium",
            expected_buckaroo_role="entity_name",
            expected_warning_type="high_uniqueness_name_not_primary_key_optional",
            should_buckaroo_warn="optional",
            profiler_failure_mode_to_test="entity_name_vs_primary_key",
            professor_question_to_answer="When should a unique name be treated as a natural key?",
            paper_claim_supported="Semantic key decisions require more than uniqueness.",
            why_this_label="Airport name names the airport entity; it may be unique but is not always a stable database key.",
            edge_case_or_risk="Names can be unique, duplicated, renamed, or formatted differently.",
        ),
        "city": location("city_name", "City is a geographic place name."),
        "state": location("state_or_region_code", "State is a regional/location code.", code=True),
        "country": location("country_name", "Country is a geographic location name."),
        "latitude": location("latitude", "Latitude is a geographic coordinate.", coordinate=True, key_like=True),
        "longitude": location("longitude", "Longitude is a geographic coordinate.", coordinate=True, key_like=True),
    }
    stocks = {
        "symbol": base_label(
            manual_true_role="identifier_code",
            manual_secondary_role="stock_ticker",
            manual_physical_type="string",
            semantic_group="entity",
            generated_label_confidence="high",
            is_foreign_key="maybe",
            should_be_key_candidate_for_buckaroo="no",
            key_rejection_reason="ticker_repeats_over_time_and_needs_date_for_composite_key",
            expected_candidate_roles="identifier_code:high; composite_key_part:high; primary_key:low",
            expected_buckaroo_role="identifier_code_composite_key_part",
            expected_warning_type="single_column_not_key_composite_possible",
            should_buckaroo_warn="yes",
            profiler_failure_mode_to_test="composite_key_part_vs_primary_key",
            professor_question_to_answer="Can Buckaroo detect that symbol plus date may identify a row, but symbol alone does not?",
            paper_claim_supported="Profilers should distinguish single-column keys from composite-key evidence.",
            why_this_label="Stock symbol identifies a traded entity, but it repeats across dates.",
            edge_case_or_risk="A sample with one date could make symbol look more key-like than it is.",
        ),
        "date": datetime_label("trading_date", "Trading date is temporal and is likely part of a symbol+date composite key.", composite=True),
        "price": money_measure("price", "stock_price", "price_measure_not_key", "Stock price is a numeric money/value measure."),
    }
    adult = {
        "age": numeric_measure("age", "integer", "bounded_numeric_demographic", "Age is a numeric demographic measure.", count=True),
        "workclass": categorical("work_class", "Workclass is a nominal employment category."),
        "fnlwgt": numeric_measure("survey_weight", "integer", "high_cardinality_weight_not_key", "Final weight is a survey/statistical weight, not identity."),
        "education": categorical("education_level", "Education label is an ordered category.", ordinal=True),
        "education.num": base_label(
            manual_true_role="ordinal_category",
            manual_secondary_role="encoded_education_level",
            manual_physical_type="integer",
            semantic_group="category",
            generated_label_confidence="high",
            is_measure_or_metric="no",
            ordinal_category="yes",
            expected_candidate_roles="ordinal_category:high; numeric_measure:medium; primary_key:low",
            expected_buckaroo_role="ordinal_category_numeric_code",
            expected_warning_type="numeric_encoded_category_optional",
            should_buckaroo_warn="optional",
            profiler_failure_mode_to_test="numeric_code_vs_measure",
            professor_question_to_answer="Can Buckaroo detect numeric-encoded categories?",
            paper_claim_supported="Candidate role output is useful when numeric values encode categories.",
            why_this_label="education.num is a numeric code for education level, not a continuous measurement.",
            edge_case_or_risk="A profiler may wrongly treat all integers as numeric measures.",
        ),
        "marital.status": categorical("marital_status", "Marital status is a nominal demographic category.", sensitive=True),
        "occupation": categorical("occupation", "Occupation is a textual job category; semantic ML may help interpret labels.", ml="maybe"),
        "relationship": categorical("household_relationship", "Relationship is a household role category.", sensitive=True),
        "race": categorical("race", "Race is a sensitive demographic category.", sensitive=True),
        "sex": categorical("sex", "Sex is a sensitive demographic category.", sensitive=True),
        "capital.gain": money_measure("capital.gain", "capital_gain", "zero_heavy_money_measure", "Capital gain is a zero-heavy money amount.", zero=True),
        "capital.loss": money_measure("capital.loss", "capital_loss", "zero_heavy_money_measure", "Capital loss is a zero-heavy money amount.", zero=True),
        "hours.per.week": numeric_measure("hours_per_week", "integer", "bounded_work_hours_measure", "Hours per week is a numeric work-time measure.", count=True),
        "native.country": location("country_name", "Native country is a location/country category."),
        "income": categorical("income_bracket", "Income is a binary ordered target category.", ordinal=True, sensitive=True),
    }
    diamonds = {
        "carat": numeric_measure("weight_measure", "float", "standard_numeric_measure", "Carat is a diamond weight measurement."),
        "cut": categorical("cut_quality", "Cut is an ordered quality category.", ordinal=True),
        "color": categorical("diamond_color_grade", "Diamond color is an ordered grading category.", ordinal=True),
        "clarity": categorical("clarity_grade", "Diamond clarity is an ordered grading category.", ordinal=True),
        "depth": numeric_measure("percentage_measure", "float", "rate_ratio_or_percentage_measure", "Depth is a percentage-like numeric measure."),
        "table": numeric_measure("percentage_measure", "float", "rate_ratio_or_percentage_measure", "Table is a percentage-like numeric measure."),
        "price": money_measure("price", "diamond_price", "price_measure_not_key", "Price is a numeric money amount."),
        "x": numeric_measure("dimension_measure", "float", "physical_dimension_measure", "X is a physical dimension measurement."),
        "y": numeric_measure("dimension_measure", "float", "physical_dimension_measure", "Y is a physical dimension measurement."),
        "z": numeric_measure("dimension_measure", "float", "physical_dimension_measure", "Z is a physical dimension measurement."),
    }
    by_dataset = {
        "taxi_trips": taxi,
        "us_airports": airports,
        "stock_prices": stocks,
        "adult_census_income": adult,
        "diamonds_pricing": diamonds,
    }
    label = by_dataset.get(dataset_id, {}).get(column)
    if label is None:
        return base_label(
            manual_true_role="unknown_or_mixed",
            generated_label_confidence="low",
            requires_semantic_ml="maybe",
            sbert_use_recommended="maybe",
            simple_rules_enough="maybe",
            why_this_label="Needs human review.",
            edge_case_or_risk="Unrecognized column name.",
        )
    return label


def add_missingness(row: Dict[str, str]) -> None:
    null_ratio = float(row.get("null_ratio") or "0")
    if null_ratio == 0:
        row["has_missing_values"] = "no"
        row["missingness_severity"] = "none"
    elif null_ratio < 0.02:
        row["has_missing_values"] = "yes"
        row["missingness_severity"] = "low"
    elif null_ratio < 0.15:
        row["has_missing_values"] = "yes"
        row["missingness_severity"] = "medium"
    else:
        row["has_missing_values"] = "yes"
        row["missingness_severity"] = "high"
    if row["has_missing_values"] == "yes" and row.get("expected_warning_type") in {"", "none"}:
        row["expected_warning_type"] = "missingness_warning"
        row["should_buckaroo_warn"] = "yes"


def fill_if_blank(row: Dict[str, str], field: str, value: str) -> None:
    if not row.get(field):
        row[field] = value


def dataset_entity_type(dataset_id: str) -> str:
    return {
        "taxi_trips": "taxi_trip_record",
        "us_airports": "airport_reference_record",
        "stock_prices": "stock_price_observation",
        "adult_census_income": "census_person_record",
        "diamonds_pricing": "diamond_product_record",
    }.get(dataset_id, "dataset_record")


def fill_range_expectations(row: Dict[str, str]) -> None:
    column = row.get("column_name", "")
    secondary = row.get("manual_secondary_role", "")
    role = row.get("manual_true_role", "")

    exact_ranges = {
        "age": ("0", "120"),
        "hours.per.week": ("0", "168"),
        "passengers": ("0", "dataset_dependent"),
        "latitude": ("-90", "90"),
        "longitude": ("-180", "180"),
        "depth": ("0", "100"),
        "table": ("0", "100"),
    }
    if column in exact_ranges:
        low, high = exact_ranges[column]
        fill_if_blank(row, "expected_min", low)
        fill_if_blank(row, "expected_max", high)
        fill_if_blank(row, "is_bounded_numeric", "yes")
    elif role == "numeric_measure":
        fill_if_blank(row, "expected_min", "0" if row.get("negative_allowed") == "no" else "dataset_dependent")
        fill_if_blank(row, "expected_max", "dataset_dependent")

    if secondary == "percentage_measure":
        row["is_rate_ratio_or_percentage"] = "yes"
        fill_if_blank(row, "is_bounded_numeric", "yes")
        fill_if_blank(row, "expected_min", "0")
        fill_if_blank(row, "expected_max", "100")
    else:
        fill_if_blank(row, "is_rate_ratio_or_percentage", "no")


def enrich_peer_review_defaults(row: Dict[str, str]) -> None:
    """Fill research metadata without overriding the hand-written labels."""
    role = row.get("manual_true_role", "")
    group = row.get("semantic_group", "")
    column = row.get("column_name", "")
    dataset_id = row.get("dataset_id", "")
    high_unique_not_key = row.get("is_high_uniqueness_but_not_key") == "yes"
    key_like = row.get("could_be_key_by_uniqueness") in {"yes", "maybe"}
    should_key = row.get("should_be_key_candidate_for_buckaroo") in {"yes", "maybe"}
    warned = row.get("should_buckaroo_warn") in {"yes", "optional"}

    fill_if_blank(row, "review_status", "needs_review")
    fill_if_blank(row, "corrected_manual_true_role", role)
    fill_if_blank(row, "corrected_is_primary_key", row.get("is_primary_key", "no"))
    fill_if_blank(row, "corrected_notes", "Auto-filled from generated label; change only if wrong.")
    fill_if_blank(row, "generated_label_confidence", "medium")
    fill_if_blank(row, "semantic_subtype", row.get("manual_secondary_role", role))
    fill_if_blank(row, "entity_type", dataset_entity_type(dataset_id))
    fill_if_blank(row, "manual_label_priority", "high" if high_unique_not_key or warned else "medium")
    fill_if_blank(row, "benchmark_importance", "high" if high_unique_not_key or role in {"identifier_code", "geographic_coordinate", "datetime"} else "medium")
    fill_if_blank(row, "poster_worthy_example", "yes" if high_unique_not_key else "maybe" if warned else "no")
    fill_if_blank(row, "is_identifier_like_code", "yes" if role == "identifier_code" else "no")

    if role == "identifier_code":
        fill_if_blank(row, "identifier_code_type", row.get("manual_secondary_role", "identifier_code"))
        fill_if_blank(row, "entity_name_vs_identifier", "identifier")
        fill_if_blank(row, "is_natural_key", "yes" if should_key else "maybe")
        fill_if_blank(row, "is_surrogate_key", "no")
    elif role == "entity_name":
        fill_if_blank(row, "entity_name_vs_identifier", "entity_name")
        fill_if_blank(row, "is_natural_key", "maybe")
        fill_if_blank(row, "is_surrogate_key", "no")
    else:
        fill_if_blank(row, "is_natural_key", "no")
        fill_if_blank(row, "is_surrogate_key", "no")
        fill_if_blank(row, "identifier_code_type", "not_applicable")
        fill_if_blank(row, "entity_name_vs_identifier", "not_applicable")

    if column == "total":
        fill_if_blank(row, "derived_or_calculated_field", "yes")
        fill_if_blank(row, "derived_from_columns", "fare + tip + tolls plus possible fees/taxes")
    else:
        fill_if_blank(row, "derived_or_calculated_field", "maybe" if role == "numeric_measure" and "total" in column else "no")
        fill_if_blank(row, "derived_from_columns", "not_applicable")

    if role == "datetime":
        fill_if_blank(row, "temporal_granularity", "timestamp" if row.get("manual_physical_type") == "datetime_string" else "date")
        fill_if_blank(row, "lifecycle_event_type", row.get("manual_secondary_role", "temporal_event"))
        fill_if_blank(row, "order_dependency_candidate", "yes")
        fill_if_blank(row, "fd_dependent_candidate", "maybe")
    elif role == "ordinal_category":
        fill_if_blank(row, "order_dependency_candidate", "yes")
        fill_if_blank(row, "fd_dependent_candidate", "yes")
    else:
        fill_if_blank(row, "order_dependency_candidate", "no")

    if group == "location":
        fill_if_blank(row, "geographic_level", row.get("manual_secondary_role", "location"))
        fill_if_blank(row, "location_semantic_type", row.get("manual_secondary_role", "location"))
        fill_if_blank(row, "domain_vocabulary", "location_reference_values")
    else:
        fill_if_blank(row, "geographic_level", "not_applicable")
        fill_if_blank(row, "location_semantic_type", "not_applicable")
    if role == "geographic_coordinate":
        fill_if_blank(row, "coordinate_pair_partner", "longitude" if column == "latitude" else "latitude" if column == "longitude" else "")
        fill_if_blank(row, "is_bounded_numeric", "yes")
        if column == "latitude":
            fill_if_blank(row, "expected_min", "-90")
            fill_if_blank(row, "expected_max", "90")
        elif column == "longitude":
            fill_if_blank(row, "expected_min", "-180")
            fill_if_blank(row, "expected_max", "180")

    if role == "numeric_measure":
        fill_if_blank(row, "is_bounded_numeric", "maybe")
        fill_if_blank(row, "negative_allowed", "maybe")
        fill_if_blank(row, "outlier_sensitive", "yes")
        fill_if_blank(row, "fd_dependent_candidate", "yes")
    if row.get("is_money_amount") == "yes":
        fill_if_blank(row, "unit_or_currency", "USD_or_dataset_currency")
        fill_if_blank(row, "negative_allowed", "usually_no")
    if row.get("is_count_or_quantity") == "yes":
        fill_if_blank(row, "expected_min", "0")
        fill_if_blank(row, "negative_allowed", "no")
    if role in {"categorical", "ordinal_category"}:
        fill_if_blank(row, "fd_dependent_candidate", "yes")
        fill_if_blank(row, "domain_vocabulary", row.get("manual_secondary_role", "dataset_specific_values"))
    elif role not in {"location_name", "geographic_coordinate"}:
        fill_if_blank(row, "domain_vocabulary", "not_applicable")
    if row.get("sensitive_or_pii_risk") == "yes":
        fill_if_blank(row, "pii_type", row.get("manual_secondary_role", "sensitive_attribute"))
    else:
        fill_if_blank(row, "pii_type", "not_applicable")

    if row.get("is_primary_key") == "yes":
        fill_if_blank(row, "ucc_candidate_status", "confirmed_single_column_key")
        fill_if_blank(row, "fd_determinant_candidate", "yes")
    elif should_key:
        fill_if_blank(row, "ucc_candidate_status", "candidate_needs_review")
        fill_if_blank(row, "fd_determinant_candidate", "maybe")
    elif key_like:
        fill_if_blank(row, "ucc_candidate_status", "statistically_unique_but_semantically_rejected")
        fill_if_blank(row, "fd_determinant_candidate", "maybe")
    else:
        fill_if_blank(row, "ucc_candidate_status", "not_a_key_candidate")
        fill_if_blank(row, "fd_determinant_candidate", "no")

    if row.get("is_foreign_key") == "yes":
        fill_if_blank(row, "ind_candidate_status", "likely_inclusion_dependency")
    elif row.get("is_foreign_key") == "maybe" or group in {"location", "entity"}:
        fill_if_blank(row, "ind_candidate_status", "possible_reference_or_lookup")
    else:
        fill_if_blank(row, "ind_candidate_status", "not_expected")

    if dataset_id == "stock_prices" and column == "symbol":
        row["is_composite_key_part"] = "yes"
        row["possible_composite_key_with"] = "date"
    elif dataset_id == "stock_prices" and column == "date":
        row["is_composite_key_part"] = "yes"
        row["possible_composite_key_with"] = "symbol"
    else:
        fill_if_blank(row, "is_composite_key_part", "no")
        fill_if_blank(row, "possible_composite_key_with", "not_applicable")

    if row.get("is_foreign_key") in {"yes", "maybe"}:
        fill_if_blank(row, "foreign_key_target_if_known", "possible_reference_table_unknown")
    elif group == "location":
        fill_if_blank(row, "foreign_key_target_if_known", "possible_location_lookup_table")
    else:
        fill_if_blank(row, "foreign_key_target_if_known", "not_applicable")

    if high_unique_not_key:
        fill_if_blank(row, "sample_size_sensitivity", "high")
        fill_if_blank(row, "small_sample_false_key_risk", "high")
        fill_if_blank(row, "min_recommended_sample_size", "10000_or_until_confidence_interval_stabilizes")
        fill_if_blank(row, "confidence_interval_priority", "high")
    elif key_like or role in {"identifier_code", "entity_name"}:
        fill_if_blank(row, "sample_size_sensitivity", "medium")
        fill_if_blank(row, "small_sample_false_key_risk", "medium")
        fill_if_blank(row, "min_recommended_sample_size", "5000_or_until_confidence_interval_stabilizes")
        fill_if_blank(row, "confidence_interval_priority", "high")
    else:
        fill_if_blank(row, "sample_size_sensitivity", "medium" if role in {"categorical", "ordinal_category"} else "low")
        fill_if_blank(row, "small_sample_false_key_risk", "low")
        fill_if_blank(row, "min_recommended_sample_size", "1000_or_confidence_interval_based")
        fill_if_blank(row, "confidence_interval_priority", "medium")

    fill_if_blank(row, "noise_sensitivity", "high" if role in {"datetime", "geographic_coordinate"} or row.get("sensitive_or_pii_risk") == "yes" else "medium")
    fill_if_blank(row, "invalid_value_risk", "yes" if row.get("is_bounded_numeric") == "yes" or role in {"datetime", "geographic_coordinate"} else "maybe")
    fill_if_blank(row, "standardization_risk", "yes" if group in {"location", "category", "entity"} else "maybe")
    fill_if_blank(row, "mixed_type_risk", "maybe")
    fill_if_blank(row, "format_consistency_risk", "yes" if role in {"datetime", "identifier_code"} else "maybe")
    fill_if_blank(row, "unit_consistency_risk", "yes" if row.get("is_measure_or_metric") == "yes" else "no")
    fill_if_blank(row, "free_text_or_description", "yes" if role == "entity_name" else "no")
    fill_if_blank(row, "boolean_like", "no")
    fill_if_blank(row, "zero_heavy", "yes" if "zero_heavy" in row.get("profiler_failure_mode_to_test", "") else "no")

    if not row.get("missingness_semantics"):
        row["missingness_semantics"] = "complete_column" if row.get("has_missing_values") == "no" else "missing_values_need_domain_review"
    if not row.get("negative_allowed"):
        row["negative_allowed"] = "no" if row.get("is_measure_or_metric") == "yes" else "not_applicable"
    if not row.get("outlier_sensitive"):
        row["outlier_sensitive"] = "yes" if row.get("is_measure_or_metric") == "yes" else "no"
    fill_if_blank(row, "coordinate_pair_partner", "not_applicable")
    fill_if_blank(row, "temporal_granularity", "not_applicable")
    fill_if_blank(row, "lifecycle_event_type", "not_applicable")
    fill_range_expectations(row)
    fill_if_blank(row, "is_bounded_numeric", "no")
    fill_if_blank(row, "expected_min", "not_applicable")
    fill_if_blank(row, "expected_max", "not_applicable")
    fill_if_blank(row, "unit_or_currency", "dataset_unit_or_unitless" if row.get("is_measure_or_metric") == "yes" else "not_applicable")
    fill_if_blank(row, "fd_dependent_candidate", "no")

    fill_if_blank(row, "expected_candidate_confidence_pattern", f"{role}:highest; alternate_roles:lower; primary_key:low_unless_semantically_supported")
    fill_if_blank(row, "low_confidence_adaptive_sampling_trigger", "sample more rows when top role confidence is low or confidence interval overlaps another role")
    fill_if_blank(row, "adaptive_sampling_stop_condition", "stop when Wilson/confidence interval is narrow enough or max sample budget is reached")
    fill_if_blank(row, "ui_should_show_confidence_interval", "yes")
    fill_if_blank(row, "ui_should_show_warning_badge", "yes" if warned else "no")
    fill_if_blank(row, "ui_user_facing_explanation", f"{column} is labeled as {role} because of its name, values, uniqueness, missingness, and domain meaning.")

    if role == "numeric_measure":
        fill_if_blank(row, "deequ_expected_constraint", "Completeness plus min/max/mean/stddev/outlier-style checks; uniqueness is not enough for identity.")
    elif role in {"categorical", "ordinal_category"}:
        fill_if_blank(row, "deequ_expected_constraint", "Completeness, distinctness, entropy, and allowed-value/category distribution checks.")
    elif role == "datetime":
        fill_if_blank(row, "deequ_expected_constraint", "Completeness, parseability, min/max date, and high-uniqueness warning checks.")
    else:
        fill_if_blank(row, "deequ_expected_constraint", "Completeness, distinctness, uniqueness, and domain-specific validity checks.")

    if should_key:
        fill_if_blank(row, "metanome_expected_behavior", "HyUCC may find this as a unique column candidate; HyFD may treat it as a determinant.")
    elif high_unique_not_key:
        fill_if_blank(row, "metanome_expected_behavior", "HyUCC may report mathematical uniqueness, but semantic Buckaroo should not promote it to primary key.")
    else:
        fill_if_blank(row, "metanome_expected_behavior", "Metanome may report dependencies/distinctness, but this is not expected to be a primary key.")
    fill_if_blank(row, "dataprofiler_expected_behavior", f"Should infer physical type plus statistics; semantic label may need domain hints for {role}.")
    fill_if_blank(row, "llm_semantic_label_expected", "Useful for explaining domain meaning; should not override statistical evidence without confidence.")
    fill_if_blank(
        row,
        "advanced_ml_analysis_reason",
        "Semantic ML may help interpret domain-specific names/values." if row.get("requires_semantic_ml") in {"yes", "maybe"} else "Simple rules, statistics, and dictionaries should be enough.",
    )
    fill_if_blank(row, "reviewer_notes", "Not reviewed yet.")

    for field in FIELDS:
        fill_if_blank(row, field, "not_applicable")


def generated_rows() -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for dataset_id, path in DATASETS:
        for evidence in dataset_stats(dataset_id, path):
            label = labels_for(dataset_id, evidence["column_name"])
            row = {field: "" for field in FIELDS}
            row.update(label)
            row.update(evidence)
            add_missingness(row)
            enrich_peer_review_defaults(row)
            rows.append(row)
    return rows


def write_generated_csv(rows: List[Dict[str, str]]) -> None:
    with GENERATED_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in FIELDS})


def parse_raw_value(dataset_id: str, header: str, value: str):
    if value == "":
        return None
    date_columns = {"taxi_trips": {"pickup", "dropoff"}, "stock_prices": {"date"}}
    text_columns = {
        "taxi_trips": {"color", "payment", "pickup_zone", "dropoff_zone", "pickup_borough", "dropoff_borough"},
        "us_airports": {"iata", "name", "city", "state", "country"},
        "stock_prices": {"symbol"},
        "adult_census_income": {
            "workclass",
            "education",
            "marital.status",
            "occupation",
            "relationship",
            "race",
            "sex",
            "native.country",
            "income",
        },
        "diamonds_pricing": {"cut", "color", "clarity"},
    }
    if header in date_columns.get(dataset_id, set()):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass
        return value
    if header not in text_columns.get(dataset_id, set()):
        try:
            return float(value) if "." in value else int(value)
        except ValueError:
            return value
    return value


def write_raw_rows(ws, rows: List[List[str]], dataset_id: str) -> None:
    headers = rows[0]
    ws.append(headers)
    for row in rows[1:]:
        ws.append([parse_raw_value(dataset_id, headers[idx], value) for idx, value in enumerate(row)])


def width_for_label(header: str) -> int:
    small = {
        "review_status",
        "row_count",
        "null_ratio",
        "unique_ratio",
        "non_null_count",
        "null_count",
        "unique_count",
        "duplicate_value_count",
        "top_value_count",
        "top_value_ratio",
        "is_primary_key",
        "is_foreign_key",
        "has_missing_values",
        "sbert_use_recommended",
        "should_buckaroo_warn",
    }
    medium = {
        "dataset_id",
        "column_name",
        "manual_true_role",
        "manual_secondary_role",
        "manual_physical_type",
        "semantic_group",
        "generated_label_confidence",
        "requires_semantic_ml",
        "simple_rules_enough",
        "adaptive_sampling_priority",
        "expected_buckaroo_role",
        "expected_warning_type",
        "manual_label_priority",
        "benchmark_importance",
        "poster_worthy_example",
        "confidence_interval_priority",
        "min_recommended_sample_size",
    }
    large = {
        "sample_values",
        "common_values",
        "top_value",
        "expected_candidate_roles",
        "expected_candidate_confidence_pattern",
        "low_confidence_adaptive_sampling_trigger",
        "adaptive_sampling_stop_condition",
        "advanced_ml_analysis_reason",
        "metanome_expected_behavior",
        "deequ_expected_constraint",
        "dataprofiler_expected_behavior",
        "llm_semantic_label_expected",
        "ui_user_facing_explanation",
        "profiler_failure_mode_to_test",
        "professor_question_to_answer",
        "paper_claim_supported",
        "why_this_label",
        "edge_case_or_risk",
        "reviewer_notes",
        "corrected_notes",
    }
    if header in small:
        return 15
    if header in medium:
        return 23
    if header in large:
        return 48
    return 24


def width_for_raw(dataset_id: str, header: str) -> int:
    if header in {"pickup", "dropoff"}:
        return 23
    if header == "date":
        return 14
    if header in {"pickup_zone", "dropoff_zone", "name", "city", "occupation", "native.country", "education"}:
        return 30
    if dataset_id == "diamonds_pricing":
        return 13
    return max(12, min(26, len(header) + 4))


INK = "12343B"
WHITE = "FFFFFF"
PALE = "F6FAFC"
CREAM = "FFFDF6"
GREEN = "E8F6EF"
LINE = "DDEAF0"
header_fill = PatternFill("solid", fgColor=INK)
evidence_fill = PatternFill("solid", fgColor=PALE)
label_fill = PatternFill("solid", fgColor=CREAM)
review_fill = PatternFill("solid", fgColor=GREEN)
header_font = Font(name="Aptos", size=10, bold=True, color=WHITE)
body_font = Font(name="Aptos", size=10, color="1F2933")
small_font = Font(name="Aptos", size=9, color="1F2933")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_alignment = Alignment(vertical="top", wrap_text=True)
thin_bottom = Border(bottom=Side(style="thin", color=LINE))


def style_header(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_bottom
    ws.row_dimensions[1].height = 46


def add_validation(ws, header_to_values: Dict[str, str]) -> None:
    headers = [cell.value for cell in ws[1]]
    for idx, header in enumerate(headers, start=1):
        if header not in header_to_values:
            continue
        dv = DataValidation(type="list", formula1=f'"{header_to_values[header]}"', allow_blank=True)
        ws.add_data_validation(dv)
        col = get_column_letter(idx)
        dv.add(f"{col}2:{col}{ws.max_row}")


def style_label_sheet(ws) -> None:
    style_header(ws)
    headers = [cell.value for cell in ws[1]]
    for idx, header in enumerate(headers, start=1):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = width_for_label(str(header))
        fill = review_fill if idx <= 4 else evidence_fill if idx <= 18 else label_fill
        for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2, max_row=ws.max_row):
            for c in cell:
                c.fill = fill
                c.font = body_font
                c.alignment = body_alignment
                c.border = thin_bottom
                if str(header).endswith("_ratio"):
                    c.number_format = "0.000"
                elif header == "row_count":
                    c.number_format = "#,##0"
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 82
    add_validation(
        ws,
        {
            "review_status": "needs_review,accepted,change_needed,unsure",
            "corrected_is_primary_key": "yes,no,maybe",
            "corrected_manual_true_role": "datetime,numeric_measure,categorical,ordinal_category,location_name,geographic_coordinate,entity_name,identifier_code,text,unknown_or_mixed",
            "manual_true_role": "datetime,numeric_measure,categorical,ordinal_category,location_name,geographic_coordinate,entity_name,identifier_code,text,unknown_or_mixed",
            "is_primary_key": "yes,no,maybe",
            "is_foreign_key": "yes,no,maybe",
            "could_be_key_by_uniqueness": "yes,no,maybe",
            "should_be_key_candidate_for_buckaroo": "yes,no,maybe",
            "is_high_uniqueness_but_not_key": "yes,no,maybe",
            "requires_semantic_ml": "yes,no,maybe",
            "sbert_use_recommended": "yes,no,maybe",
            "simple_rules_enough": "yes,no,maybe",
            "adaptive_sampling_priority": "low,medium,high",
            "generated_label_confidence": "low,medium,high",
            "benchmark_importance": "low,medium,high",
            "manual_label_priority": "low,medium,high",
            "poster_worthy_example": "yes,no,maybe",
            "ui_should_show_confidence_interval": "yes,no",
            "ui_should_show_warning_badge": "yes,no",
            "should_buckaroo_warn": "yes,no,optional",
        },
    )


def style_raw_sheet(ws, dataset_id: str) -> None:
    style_header(ws)
    headers = [cell.value for cell in ws[1]]
    ws.sheet_format.defaultRowHeight = 24
    for idx, header in enumerate(headers, start=1):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = width_for_raw(dataset_id, str(header))
        if header in {"pickup", "dropoff", "date"}:
            fmt = "yyyy-mm-dd hh:mm:ss" if header in {"pickup", "dropoff"} else "yyyy-mm-dd"
            for row_idx in range(2, ws.max_row + 1):
                ws[f"{col}{row_idx}"].number_format = fmt


def add_codebook(wb: Workbook) -> None:
    ws = wb.create_sheet("Codebook")
    rows = [("Field", "Group", "Meaning", "How to review it")]
    for field in FIELDS:
        rows.append((field, codebook_group(field), codebook_meaning(field), codebook_review_hint(field)))
    for row in rows:
        ws.append(row)
    style_header(ws)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 78
    ws.column_dimensions["D"].width = 70
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = label_fill
            cell.font = body_font
            cell.alignment = body_alignment
            cell.border = thin_bottom
        ws.row_dimensions[row[0].row].height = 58


def codebook_group(field: str) -> str:
    if field.startswith("corrected_") or field == "review_status" or field == "reviewer_notes":
        return "review"
    if field in {
        "dataset_id",
        "column_name",
        "row_count",
        "null_ratio",
        "unique_ratio",
        "sample_values",
        "common_values",
        "non_null_count",
        "null_count",
        "unique_count",
        "duplicate_value_count",
        "top_value",
        "top_value_count",
        "top_value_ratio",
    }:
        return "evidence"
    if field.startswith("manual_") or field in {
        "semantic_group",
        "semantic_subtype",
        "unit_or_currency",
        "temporal_granularity",
        "geographic_level",
        "entity_type",
        "domain_vocabulary",
        "generated_label_confidence",
    }:
        return "semantic label"
    if "key" in field or field in {
        "is_identifier_like_code",
        "identifier_code_type",
        "entity_name_vs_identifier",
        "is_surrogate_key",
        "is_natural_key",
        "is_composite_key_part",
        "possible_composite_key_with",
        "foreign_key_target_if_known",
    }:
        return "key / identity"
    if field.startswith("is_") or field.endswith("_risk") or field.startswith("has_") or field.startswith("missingness"):
        return "edge case / quality"
    if field in {"ucc_candidate_status", "fd_determinant_candidate", "fd_dependent_candidate", "ind_candidate_status", "order_dependency_candidate"}:
        return "dependency"
    if "semantic_ml" in field or "sbert" in field or "model" in field or "rules_enough" in field:
        return "ML decision"
    if "sampling" in field or "confidence_interval" in field or "sample_size" in field:
        return "sampling / confidence"
    if field.startswith("expected_") or field.startswith("ui_") or field.startswith("profiler_"):
        return "Buckaroo expectation"
    if "metanome" in field or "deequ" in field or "dataprofiler" in field or "llm" in field:
        return "baseline comparison"
    return "research notes"


def codebook_meaning(field: str) -> str:
    meanings = {
        "review_status": "Your review decision for the generated label.",
        "corrected_manual_true_role": "Use this only if you disagree with the generated true semantic role.",
        "corrected_is_primary_key": "Use this only if the generated key decision is wrong or uncertain.",
        "corrected_notes": "Short explanation for any correction you make.",
        "dataset_id": "Dataset the column came from.",
        "column_name": "Original column name in the CSV.",
        "row_count": "Number of rows in the source CSV.",
        "null_ratio": "Share of rows where this column is blank.",
        "unique_ratio": "Distinct non-empty values divided by non-empty rows.",
        "sample_values": "A few example values from the column.",
        "common_values": "Most frequent values and their counts.",
        "non_null_count": "Number of non-empty values.",
        "null_count": "Number of blank values.",
        "unique_count": "Number of distinct non-empty values.",
        "duplicate_value_count": "Non-empty rows minus distinct values; high means values repeat.",
        "top_value": "Most common value in the column.",
        "top_value_count": "How many times the most common value appears.",
        "top_value_ratio": "Top value count divided by non-empty row count.",
        "manual_true_role": "Generated human semantic truth label, such as datetime, numeric_measure, categorical, or identifier_code.",
        "manual_secondary_role": "More specific meaning inside the main role.",
        "manual_physical_type": "Surface data type: string, integer, float, datetime string, etc.",
        "semantic_group": "Broad domain group: temporal, measure, location, category, entity, etc.",
        "semantic_subtype": "Optional deeper semantic subtype.",
        "unit_or_currency": "Known unit or currency, if the values are measurements or money.",
        "temporal_granularity": "Whether a time column is a date, timestamp, month, year, etc.",
        "geographic_level": "City, country, coordinate, airport code, zone, borough, or another place level.",
        "entity_type": "The real-world object described by the column, if relevant.",
        "domain_vocabulary": "Expected controlled value set or domain dictionary.",
        "generated_label_confidence": "How confident the generated human label is.",
        "is_primary_key": "Whether this column truly identifies each row by itself.",
        "is_foreign_key": "Whether this column likely references another table.",
        "could_be_key_by_uniqueness": "Whether uniqueness statistics alone could tempt a profiler to call it a key.",
        "should_be_key_candidate_for_buckaroo": "Whether Buckaroo should seriously present this as a key candidate.",
        "is_high_uniqueness_but_not_key": "Important false-key flag: many unique values, but not row identity.",
        "key_rejection_reason": "Why a key-looking column should not be treated as a true key.",
        "requires_semantic_ml": "Whether rules/statistics are insufficient and an ML/LLM semantic model could help.",
        "recommended_semantic_model": "Suggested method: rules, dictionary, SBERT, LLM, parser, etc.",
        "sbert_use_recommended": "Whether sentence-embedding style analysis is worth using.",
        "simple_rules_enough": "Whether simple rules/statistics should be enough.",
        "adaptive_sampling_priority": "How strongly Buckaroo should sample more rows before finalizing this column.",
        "expected_buckaroo_role": "Role the improved Buckaroo profiler should output.",
        "expected_warning_type": "Warning Buckaroo should show, such as unique timestamp not primary key.",
        "should_buckaroo_warn": "Whether the UI should show a warning for this column.",
        "profiler_failure_mode_to_test": "The exact kind of profiling mistake this column is useful for testing.",
        "paper_claim_supported": "Research claim this label helps support.",
    }
    if field in meanings:
        return meanings[field]
    if field.startswith("is_"):
        return "Yes/no/maybe marker for a semantic or data-quality property."
    if field.endswith("_risk"):
        return "Risk label showing where a profiler may need safeguards or warnings."
    if field.startswith("expected_"):
        return "Expected behavior from Buckaroo or a comparison profiler."
    if field.endswith("_candidate") or field.endswith("_candidate_status"):
        return "Candidate status for a key, dependency, or relationship discovery algorithm."
    return "Supporting research annotation used to compare profiler behavior and explain edge cases."


def codebook_review_hint(field: str) -> str:
    if field.startswith("corrected_"):
        return "Leave blank unless you need to override the generated label."
    if field == "review_status":
        return "Change to accepted if the row looks correct, change_needed if not, unsure if you need help."
    if field in {"manual_true_role", "is_primary_key", "should_be_key_candidate_for_buckaroo", "is_high_uniqueness_but_not_key"}:
        return "Review this carefully; these fields directly affect profiler accuracy scoring."
    if field in {"requires_semantic_ml", "sbert_use_recommended", "simple_rules_enough"}:
        return "Check whether a human can understand the column from name/values, or whether semantic ML would add value."
    if field in {"sample_values", "common_values", "top_value"}:
        return "Use this as evidence; do not edit unless the generated evidence is wrong."
    return "Usually review only; edit if the generated label does not match your understanding."


def add_dataset_index(wb: Workbook, rows: List[Dict[str, str]]) -> None:
    ws = wb.create_sheet("Dataset Index")
    ws.append(["dataset_id", "topic", "source_rows", "columns_labeled", "why_this_dataset_matters"])
    counts = Counter(row["dataset_id"] for row in rows)
    row_counts = {row["dataset_id"]: row["row_count"] for row in rows}
    for dataset_id, _path in DATASETS:
        topic, why = DATASET_CONTEXT[dataset_id]
        ws.append([dataset_id, topic, int(row_counts.get(dataset_id, "0")), counts[dataset_id], why])
    style_header(ws)
    widths = [26, 30, 16, 18, 90]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = evidence_fill
            cell.font = body_font
            cell.alignment = body_alignment
            cell.border = thin_bottom
        ws.row_dimensions[row[0].row].height = 52


def add_review_guide(wb: Workbook) -> None:
    ws = wb.create_sheet("Review Guide")
    ws.append(["Section", "Easy explanation"])
    for section, explanation in REVIEW_GUIDE_SECTIONS:
        ws.append([section, explanation])
    style_header(ws)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 118
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = label_fill
            cell.font = body_font
            cell.alignment = body_alignment
            cell.border = thin_bottom
        ws.row_dimensions[row[0].row].height = 72


def write_review_guide_md() -> None:
    lines = [
        "# Manual Label Review Guide",
        "",
        "Use this guide while reviewing `manual_labeling_peer_review_final.xlsx`.",
        "",
    ]
    for section, explanation in REVIEW_GUIDE_SECTIONS:
        lines.extend([f"## {section}", "", explanation, ""])
    REVIEW_GUIDE.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    rows = generated_rows()
    write_generated_csv(rows)
    write_review_guide_md()

    wb = Workbook()
    wb.remove(wb.active)

    add_dataset_index(wb, rows)
    add_review_guide(wb)

    ws = wb.create_sheet("Generated Label Review")
    ws.append(FIELDS)
    for row in rows:
        ws.append([row.get(field, "") for field in FIELDS])
    style_label_sheet(ws)

    for dataset_id, path in DATASETS:
        ws = wb.create_sheet(dataset_id)
        write_raw_rows(ws, read_csv(path), dataset_id)
        style_raw_sheet(ws, dataset_id)

    add_codebook(wb)
    saved_output = OUTPUT
    try:
        wb.save(saved_output)
    except PermissionError:
        saved_output = FILLED_OUTPUT
        wb.save(saved_output)

    check = load_workbook(saved_output, read_only=True, data_only=True)
    print(saved_output)
    print(GENERATED_CSV)
    print(REVIEW_GUIDE)
    print(f"sheets={len(check.sheetnames)}")
    for name in check.sheetnames:
        ws = check[name]
        print(f"{name}: rows={ws.max_row} cols={ws.max_column}")
    check.close()


if __name__ == "__main__":
    main()
